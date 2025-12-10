"""
Mela SIM Sell Portal - Flask Backend API
Database: Oracle 19c (dwhdb02)
Developed for: Teletalk Mela SIM Sell CRM Portal
"""

from flask import Flask, request, jsonify, send_from_directory, send_file
from flask_cors import CORS
import oracledb
import os
from datetime import datetime
import logging
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

app = Flask(__name__, static_folder='static')
CORS(app)

# Create logs directory if it doesn't exist
os.makedirs('logs', exist_ok=True)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(levelname)s | %(message)s',
    handlers=[
        logging.FileHandler('logs/api.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Database configuration (same format as Call Drop Report Portal)
DB_CONFIG = {
    'user': 'your_username',
    'password': 'your_password',
    'host': 'your_host',
    'port': 1521,
    'sid': 'your_sid',
}

# Load fixed reference data from JSON
REFERENCE_DATA_FILE = 'reference_data.json'

def load_reference_data():
    """Load fixed zones and officers from reference_data.json"""
    try:
        with open(REFERENCE_DATA_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
            logger.info(f"Loaded {len(data['zones'])} zones and {len(data['field_officers'])} field officers")
            return data
    except Exception as e:
        logger.error(f"Error loading reference data: {e}")
        return {'zones': [], 'field_officers': []}


def get_db_connection():
    """Create and return database connection using oracledb"""
    try:
        dsn = f"(DESCRIPTION=(ADDRESS=(PROTOCOL=TCP)(HOST={DB_CONFIG['host']})(PORT={DB_CONFIG['port']}))(CONNECT_DATA=(SID={DB_CONFIG['sid']})))"
        connection = oracledb.connect(
            user=DB_CONFIG['user'],
            password=DB_CONFIG['password'],
            dsn=dsn
        )
        logger.info("Database connection established successfully")
        return connection
    except Exception as error:
        logger.error(f"Database connection error: {error}")
        raise


@app.route('/')
def index():
    """Serve the main HTML page"""
    return send_from_directory('static', 'index.html')


@app.route('/api/zones', methods=['GET'])
def get_zones():
    """Get fixed zones from reference data"""
    try:
        reference_data = load_reference_data()
        zones = reference_data.get('zones', [])
        logger.info(f"Retrieved {len(zones)} zones from reference data")
        return jsonify({'success': True, 'data': zones})
    except Exception as e:
        logger.error(f"Error fetching zones: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/field-officers', methods=['GET'])
def get_field_officers():
    """Get fixed field officers from reference data"""
    try:
        reference_data = load_reference_data()
        officers = reference_data.get('field_officers', [])
        logger.info(f"Retrieved {len(officers)} field officers from reference data")
        return jsonify({'success': True, 'data': officers})
    except Exception as e:
        logger.error(f"Error fetching field officers: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/check-msisdn/<msisdn>', methods=['GET'])
def check_msisdn(msisdn):
    """Check if MSISDN already exists in database"""
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        
        query = "SELECT COUNT(*) FROM Mela_SIM_sell_crm_cnl_T WHERE MSISDN = :msisdn"
        cursor.execute(query, {'msisdn': msisdn})
        exists = cursor.fetchone()[0] > 0
        
        cursor.close()
        connection.close()
        
        return jsonify({'success': True, 'exists': exists})
    except Exception as e:
        logger.error(f"Error checking MSISDN: {e}")
        return jsonify({'success': True, 'exists': False})


@app.route('/api/submit', methods=['POST'])
def submit_data():
    """Insert new record into database or JSON file"""
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['zone', 'field_officer', 'btsid', 'msisdn', 
                          'entry_date', 'new_sim', 'replace', 'new_retailer_count']
        
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({
                    'success': False, 
                    'error': f'Missing required field: {field}'
                }), 400
        
        # Validate MSISDN length
        if len(data['msisdn']) != 10:
            return jsonify({
                'success': False, 
                'error': 'MSISDN must be exactly 10 digits'
            }), 400
        
        # Validate MSISDN starts with 15
        if not data['msisdn'].startswith('15'):
            return jsonify({
                'success': False, 
                'error': 'MSISDN must start with 15'
            }), 400
        
        # Validate new_retailer_count range
        retailer_count = int(data['new_retailer_count'])
        if retailer_count < 0 or retailer_count > 99:
            return jsonify({
                'success': False, 
                'error': 'New retailer count must be between 0 and 99'
            }), 400
        
        # Check if MSISDN already exists in database
        connection = get_db_connection()
        cursor = connection.cursor()
        
        check_query = """
            SELECT COUNT(*) FROM Mela_SIM_sell_crm_cnl_T 
            WHERE MSISDN = :msisdn
        """
        cursor.execute(check_query, {'msisdn': data['msisdn']})
        exists = cursor.fetchone()[0] > 0
        
        if exists:
            cursor.close()
            connection.close()
            logger.warning(f"MSISDN already exists: {data['msisdn']}")
            return jsonify({
                'success': False, 
                'error': f'MSISDN {data["msisdn"]} already exists in the database'
            }), 400
        
        # Insert directly into Oracle database
        
        insert_query = """
            INSERT INTO Mela_SIM_sell_crm_cnl_T 
            (ZONE, FIELD_OFFICER, BTSID, MSISDN, ENTRY_DATE, NEW_SIM, REPLACE_SIM, 
             NEW_RETAILER_COUNT, CREATED_BY)
            VALUES (:zone, :officer, :btsid, :msisdn, TO_DATE(:entry_date, 'YYYY-MM-DD'), 
                    :new_sim, :replace, :retailer_count, :created_by)
        """
        
        cursor.execute(insert_query, {
            'zone': data['zone'],
            'officer': data['field_officer'],
            'btsid': data['btsid'],
            'msisdn': data['msisdn'],
            'entry_date': data['entry_date'],
            'new_sim': data['new_sim'],
            'replace': data['replace'],
            'retailer_count': retailer_count,
            'created_by': data.get('created_by', 'WEB_PORTAL')
        })
        
        connection.commit()
        cursor.close()
        connection.close()
        
        logger.info(f"Record inserted successfully: MSISDN={data['msisdn']}")
        return jsonify({
            'success': True, 
            'message': 'Data submitted successfully'
        })
        
    except Exception as e:
        logger.error(f"Error submitting data: {e}")
        return jsonify({
            'success': False, 
            'error': str(e)
        }), 500


@app.route('/api/records', methods=['GET'])
def get_records():
    """Get recent records from Oracle database"""
    try:
        limit = request.args.get('limit', 5, type=int)
        
        connection = get_db_connection()
        cursor = connection.cursor()
        
        query = """
            SELECT ID, ZONE, FIELD_OFFICER, BTSID, MSISDN, 
                   TO_CHAR(ENTRY_DATE, 'YYYY-MM-DD') as ENTRY_DATE,
                   NEW_SIM, REPLACE_SIM, NEW_RETAILER_COUNT, 
                   TO_CHAR(CREATED_DATE, 'YYYY-MM-DD HH24:MI:SS') as CREATED_DATE
            FROM Mela_SIM_sell_crm_cnl_T 
            ORDER BY CREATED_DATE DESC 
            FETCH FIRST :limit ROWS ONLY
        """
        
        cursor.execute(query, {'limit': limit})
        
        columns = [col[0] for col in cursor.description]
        records = []
        
        for row in cursor.fetchall():
            records.append(dict(zip(columns, row)))
        
        cursor.close()
        connection.close()
        
        return jsonify({'success': True, 'data': records})
        
    except Exception as e:
        logger.error(f"Error fetching records: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/download-template', methods=['GET'])
def download_template():
    """Generate and download Excel template"""
    try:
        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mela SIM Upload"
        
        # Define headers with date format notation
        headers = ['Zone', 'Field Officer', 'BTS ID', 'MSISDN', 
                   'Date (YYYY-MM-DD)', 'New SIM', 'Replace', 'New Retailer Count']
        
        # Style for headers
        header_fill = PatternFill(start_color="7FB560", end_color="7FB560", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Adjust column widths
        column_widths = [15, 20, 12, 15, 20, 10, 10, 18]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='Mela_SIM_Upload_Template.xlsx'
        )
    
    except Exception as e:
        logger.error(f"Error generating template: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/download-records', methods=['GET'])
def download_records():
    """Download all records as Excel file"""
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        
        query = """
            SELECT ID, ZONE, FIELD_OFFICER, BTSID, MSISDN, 
                   TO_CHAR(ENTRY_DATE, 'YYYY-MM-DD') as ENTRY_DATE,
                   NEW_SIM, REPLACE_SIM, NEW_RETAILER_COUNT, 
                   TO_CHAR(CREATED_DATE, 'YYYY-MM-DD HH24:MI:SS') as CREATED_DATE
            FROM Mela_SIM_sell_crm_cnl_T 
            ORDER BY CREATED_DATE DESC
        """
        
        cursor.execute(query)
        records = cursor.fetchall()
        columns = [col[0] for col in cursor.description]
        
        cursor.close()
        connection.close()
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mela SIM Records"
        
        # Style for headers
        header_fill = PatternFill(start_color="7FB560", end_color="7FB560", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col_num, header in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Write data
        for row_num, record in enumerate(records, 2):
            for col_num, value in enumerate(record, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Adjust column widths
        column_widths = [8, 15, 20, 12, 15, 15, 10, 10, 18, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'Mela_SIM_Records_{timestamp}.xlsx'
        )
    
    except Exception as e:
        logger.error(f"Error downloading records: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/bulk-upload', methods=['POST'])
def bulk_upload():
    """Handle bulk upload from Excel file"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'error': 'Invalid file format. Please upload Excel file'}), 400
        
        # Read Excel file
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        connection = get_db_connection()
        cursor = connection.cursor()
        
        inserted_count = 0
        errors = []
        
        # Skip header row (row 1)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Skip empty rows
                if not any(row):
                    continue
                
                zone, field_officer, btsid, msisdn, entry_date, new_sim, replace_sim, retailer_count = row
                
                # Validate MSISDN
                msisdn_str = str(msisdn).strip()
                if len(msisdn_str) != 10 or not msisdn_str.startswith('15'):
                    errors.append(f"Row {row_num}: Invalid MSISDN '{msisdn_str}'")
                    continue
                
                # Validate and normalize date format
                if isinstance(entry_date, str):
                    entry_date_str = entry_date.strip()
                    date_obj = None
                    
                    # Try multiple date formats
                    date_formats = [
                        '%Y-%m-%d',     # 2025-10-15
                        '%Y/%m/%d',     # 2025/10/15
                        '%d-%m-%Y',     # 15-10-2025
                        '%d/%m/%Y',     # 15/10/2025
                        '%m-%d-%Y',     # 10-15-2025
                        '%m/%d/%Y',     # 10/15/2025
                    ]
                    
                    for fmt in date_formats:
                        try:
                            date_obj = datetime.strptime(entry_date_str, fmt)
                            break
                        except ValueError:
                            continue
                    
                    if date_obj:
                        entry_date = date_obj.strftime('%Y-%m-%d')
                    else:
                        errors.append(f"Row {row_num}: Invalid date format '{entry_date_str}'. Supported formats: YYYY-MM-DD, YYYY/MM/DD, DD-MM-YYYY, DD/MM/YYYY")
                        continue
                        
                elif hasattr(entry_date, 'strftime'):
                    # Excel date object
                    entry_date = entry_date.strftime('%Y-%m-%d')
                else:
                    errors.append(f"Row {row_num}: Invalid date value")
                    continue
                
                # Check if MSISDN already exists
                cursor.execute("SELECT COUNT(*) FROM Mela_SIM_sell_crm_cnl_T WHERE MSISDN = :msisdn", 
                             {'msisdn': msisdn_str})
                if cursor.fetchone()[0] > 0:
                    errors.append(f"Row {row_num}: MSISDN '{msisdn_str}' already exists")
                    continue
                
                # Insert record
                insert_query = """
                    INSERT INTO Mela_SIM_sell_crm_cnl_T 
                    (ZONE, FIELD_OFFICER, BTSID, MSISDN, ENTRY_DATE, NEW_SIM, 
                     REPLACE_SIM, NEW_RETAILER_COUNT, CREATED_BY)
                    VALUES (:zone, :officer, :btsid, :msisdn, TO_DATE(:entry_date, 'YYYY-MM-DD'), 
                            :new_sim, :replace, :retailer_count, :created_by)
                """
                
                cursor.execute(insert_query, {
                    'zone': str(zone).strip(),
                    'officer': str(field_officer).strip(),
                    'btsid': str(btsid).strip(),
                    'msisdn': msisdn_str,
                    'entry_date': entry_date,
                    'new_sim': str(new_sim).strip().upper(),
                    'replace': str(replace_sim).strip().upper(),
                    'retailer_count': int(retailer_count),
                    'created_by': 'BULK_UPLOAD'
                })
                
                inserted_count += 1
                
            except Exception as row_error:
                errors.append(f"Row {row_num}: {str(row_error)}")
                continue
        
        connection.commit()
        cursor.close()
        connection.close()
        
        response_data = {
            'success': True,
            'inserted': inserted_count,
            'errors': errors if errors else None
        }
        
        logger.info(f"Bulk upload completed: {inserted_count} records inserted, {len(errors)} errors")
        return jsonify(response_data)
        
    except Exception as e:
        logger.error(f"Error in bulk upload: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    try:
        connection = get_db_connection()
        connection.close()
        
        return jsonify({
            'status': 'healthy',
            'database': 'connected',
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        return jsonify({
            'status': 'unhealthy',
            'database': 'disconnected',
            'error': str(e),
            'timestamp': datetime.now().isoformat()
        }), 500


if __name__ == '__main__':
    # Run the application
    logger.info("Starting Mela SIM Sell Portal - Oracle Database Mode")
    app.run(host='0.0.0.0', port=7010, debug=True)
